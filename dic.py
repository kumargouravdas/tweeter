# -*- coding: utf-8 -*-
"""
Created on Wed Nov 15 13:22:39 2017

@author: Swarup
"""

''' '''
import openpyxl
wb3=openpyxl.load_workbook('D:\\w\\a.xlsx')
ws3=wb3.active
wb=openpyxl.load_workbook('D:\\w\\ESSAY.xlsx')
ws=wb.active
r1=ws.max_row
wb1=openpyxl.load_workbook('D:\\w\\LIWCEXCEL.xlsx')
ws1=wb1.active
r2=ws1.max_row
print(r2)
k=1
for i in range(2,(r1+1)):
    if ws.cell(row=i,column=3).value=='y' and ws.cell(row=i,column=4).value=='n' and ws.cell(row=i,column=5).value=='n' and ws.cell(row=i,column=6).value=='n' and ws.cell(row=i,column=7).value=='n':
        dict={"PREPOSITION":0,"NUMBER":0,"AFFECT":0,"POSEMO":0,"POSFEEL":0,"OPTIM":0,"NEGEMO":0,"ANX":0,
              "ANGER":0,"SAD":0,"PRONOUN":0,"COGMECH":0,"CAUSE":0,"INSIGHT":0,"DISCREP":0,"INHIB":0,"TENTAT":0,
              "CERTAIN":0,"SENSES":0,"SEE":0,"HEAR":0,"I":0,"FEEL":0,"SOCIAL":0,"COMM":0,"OTHREF":0,"FRIENDS":0,
              "FAMILY":0,"HUMANS":0,"TIME":0,"PAST":0,"PRESENT":0,"WE":0,"FUTURE":0,"SPACE":0,"UP":0,"DOWN":0,
              "INCL":0,"EXCL":0,"MOTION":0,"OCCUP":0,"SCHOOL":0,"JOB":0,"SELF":0,"ACHEIVE":0,"LEISURE":0,"HOME":0,
              "SPORTS":0,"TV":0,"MUSIC":0,"MONEY":0,"METAPH":0,"RELIG":0,"DEATH":0,"YOU":0,"PHYSICAL":0,"BODY":0,
              "SEXUAL":0,"EATING":0,"SLEEP":0,"GROOM":0,"SWEAR":0,"NONFL":0,"FILLERS":0,"OTHER":0,"NEGATE":0,"ASSENT":0,"ARTICLE":0}
        print(ws.cell(row=i,column=1).value)
        
        m=(ws.cell(row=i,column=2).value)
        text=str(m)
        l=text.lower()
        t=l.replace("_"," ")
        t1=t.replace("-"," ")
        t2=t1.replace(","," ")
        t3=t2.replace("."," ")
        t4=t3.replace("("," ")
        t5=t4.replace(")"," ")
        t6=t5.replace("?"," ")
        t7=t6.split(" ")
        t8=list(filter(str.strip,t7))
        print(t8)
        c=1
        for x in t8:
            for j in range(1,r2):
                if x==(ws1.cell(row=j,column=1).value):
                    for li in dict:
                        if li==(ws1.cell(row=j,column=2).value):
                            print(x,li)
                            dict[li]+=1
                        else:
                            dict[li]+=0
        print(dict)
        for item in dict.values():
            ws3.cell(row=k,column=c,value=item)
            c+=1
        k+=1
wb3.save('D:\\w\\a.xlsx')            
                    
   
    
    
  



              


            
